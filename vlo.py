import openpyxl

etim = openpyxl.load_workbook('testeretim.xlsx')
master = openpyxl.load_workbook('kabelmaster+EC003248.xlsx')

etim_sheet = etim['Tekstgen']
master_sheet = master['Sheet1']
maxcol= etim_sheet.max_column


for i in etim_sheet.iter_rows():
    id =i[1].value
    for j in master_sheet.iter_rows():
        row_number =j[1].row
        if j[0].value == id:
            master_sheet.cell(row= row_number, column=7).value = i[3].value
            

master.save('kabelmaster_3248_3249.xlsx')
