import openpyxl

etim = openpyxl.load_workbook('3249_2.xlsx')
master = openpyxl.load_workbook('kabelmaster.xlsx')

etim_sheet = etim['Tekstgen']
master_sheet = master['Sheet1']
maxcol= etim_sheet.max_column

print(maxcol)
for i in etim_sheet.iter_rows():
    id =i[1].value
    row_number =i[1].row
    for j in master_sheet.iter_rows():
        if j[0].value == id:
            etim_sheet.cell(row=row_number, column=maxcol + 1).value = j[5].value


etim.save('testeretim.xlsx')