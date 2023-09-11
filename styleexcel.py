from openpyxl import load_workbook
from openpyxl.styles import PatternFillFill
from openpyxl.styles import Font

def loadworkbook(filename):
    load_workbook(filename= 'f{filename}.xlsx')



def setworksheet(filename):
    wb = load_workbook(filename = 'f"{filename}".xlsx')
    worksheet = wb.active
    worksheet1 = wb.create_sheet(f"{filename}",0)
    worksheet2 = wb.create_sheet("kenmerken",1)
    worksheet3 = wb.create_sheet("synoniemen",2)
    
    




def setheader(wb):
    ws = wb.active
    ft = Font(bold=True)
    data = ["Bestelnummer","Merk","Serie","Type", "Artikelnummer Fabrikant", "Etim Klasse", "Orginele webtitel"]
    for row in data:
        ws.append(row)

    ws['A1']



def getvalue():
    for row in ws.value:
        for value in row:
            ...




def savetofile(file):
    wb = Workbook()
    wb.save(f"{file}".xlsx)

