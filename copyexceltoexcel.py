import openpyxl as xl
import pandas as pd
from getfiles import getfiles;

def copyexcel(file):
    filename = f'{file}.xlsx'

    wb1 = xl.load_workbook(filename)
    ws1 = wb1.worksheets[0]

    filename1 = 'kabelmaster.xlsx'
    wb2 = xl.load_workbook(filename1)
    ws = wb2.worksheets[0]
    
    
    

    mr = ws1.max_row
    mc = ws1.max_column

    maxrow = 1

    for i in range(1,mr + 1):
        for j in range(1, mc + 1):
            c = ws1.cell(row = i, column= j)
            ws.cell(row = maxrow, column = j).value = c.value
        maxrow = maxrow + 1
    wb2.save(str(filename1))


def main():
    files = getfiles('EtimKlasse_kabel')
    i = 0
    while i < len(files):
        copyexcel(files[i])
        
        i= i + 1
    
    

main()
